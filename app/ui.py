import os
import sys

import pandas as pd
import pyodbc
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QProgressBar, QDateEdit,
    QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox,
    QGroupBox, QGridLayout, QSplitter, QFrame, QHeaderView,
    QTextEdit, QTabWidget, QCheckBox, QSpinBox, QStackedWidget, QComboBox,
    QListWidget, QListWidgetItem, QMenu, QWidgetAction, QScrollArea
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt5.QtGui import QFont

from .styles import STYLE
from .utils import safe_filename, prepare_df_for_excel
from .workers import WorkerThread, BatchWorkerThread
from .widgets import MultiSelectButton

"""Main desktop window and startup logic."""

class InitWorker(QThread):
    """Load filter values used by the GUI at startup."""
    companies_loaded = pyqtSignal(list)   # список строк компаний
    types_loaded     = pyqtSignal(list)   # список типов вагонов
    error            = pyqtSignal(str)

    def __init__(self, conn_str):
        super().__init__()
        self.conn_str = conn_str

    def run(self):
        try:
            conn = pyodbc.connect(self.conn_str)

            df_comp = pd.read_sql("""
                SELECT DISTINCT Men.Name AS В_управлении
                FROM dbo.wagon_ownership_history r
                LEFT JOIN dbo.counterparties_reference Men ON r.ID_MANAGER = Men.ID
                WHERE Men.Name IS NOT NULL AND Men.Name <> ''
                ORDER BY Men.Name
            """, conn)
            companies = sorted(df_comp['В_управлении'].dropna().astype(str).tolist())
            self.companies_loaded.emit(companies)

            try:
                df_types = pd.read_sql(
                    "SELECT DISTINCT Name FROM dbo.wagon_type_reference WHERE Name IS NOT NULL AND Name <> '' ORDER BY Name",
                    conn)
                types = sorted(df_types['Name'].dropna().astype(str).tolist())
                self.types_loaded.emit(types)
            except Exception:
                try:
                    df_types = pd.read_sql(
                        "SELECT DISTINCT ROsOwners_CarTypeName AS Name FROM dbo.wagon_type_history WHERE ROsOwners_CarTypeName IS NOT NULL AND ROsOwners_CarTypeName <> '' ORDER BY Name",
                        conn)
                    types = sorted(df_types['Name'].dropna().astype(str).tolist())
                    self.types_loaded.emit(types)
                except Exception:
                    pass

            conn.close()
        except Exception as e:
            self.error.emit(str(e))

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.result_df = None
        self.worker    = None
        self.setWindowTitle("Railcar Turnover Analyzer")
        self.setMinimumSize(1280, 820)
        self.setStyleSheet(STYLE)
        self._build_ui()
        self._load_companies()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(12)

        hdr = QHBoxLayout()
        icon_lbl = QLabel("🚂")
        icon_lbl.setFont(QFont("Segoe UI Emoji", 28))
        title_col = QVBoxLayout()
        title_lbl = QLabel("Оборот вагонов")
        title_lbl.setObjectName("titleLabel")
        sub_lbl = QLabel("Desktop reporting tool for wagon turnover analytics")
        sub_lbl.setObjectName("subtitleLabel")
        title_col.addWidget(title_lbl)
        title_col.addWidget(sub_lbl)
        hdr.addWidget(icon_lbl)
        hdr.addSpacing(10)
        hdr.addLayout(title_col)
        hdr.addStretch()
        root.addLayout(hdr)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color: #2d3748;")
        root.addWidget(line)

        settings_row = QHBoxLayout()
        settings_row.setSpacing(14)

        conn_box  = QGroupBox("🔌 Подключение к БД")
        conn_grid = QGridLayout(conn_box)
        conn_grid.setSpacing(8)
        self.f_server   = self._field("YOUR_SERVER")
        self.f_database = self._field("YOUR_DATABASE")
        self.f_username = self._field("YOUR_USERNAME")
        self.f_password = self._field("", password=True)
        for row, (lbl, fld) in enumerate([
            ("Сервер",  self.f_server),
            ("База",    self.f_database),
            ("Логин",   self.f_username),
            ("Пароль",  self.f_password),
        ]):
            conn_grid.addWidget(QLabel(lbl), row, 0)
            conn_grid.addWidget(fld,          row, 1)
        settings_row.addWidget(conn_box, 3)

        period_outer = QGroupBox("📅 Период")
        period_outer_layout = QVBoxLayout(period_outer)
        period_outer_layout.setSpacing(6)

        mode_row = QHBoxLayout()
        self.chk_single = QCheckBox("Одиночный")
        self.chk_single.setChecked(True)
        self.chk_batch  = QCheckBox("Пакетный (серия файлов)")
        self.chk_single.setStyleSheet("color:#90cdf4; font-weight:600;")
        self.chk_batch.setStyleSheet("color:#68d391; font-weight:600;")
        mode_row.addWidget(self.chk_single)
        mode_row.addSpacing(16)
        mode_row.addWidget(self.chk_batch)
        mode_row.addStretch()
        period_outer_layout.addLayout(mode_row)

        self.period_stack = QStackedWidget()

        page_single = QWidget()
        pg_grid = QGridLayout(page_single)
        pg_grid.setSpacing(8)
        pg_grid.setContentsMargins(0, 0, 0, 0)
        self.f_date_from = QDateEdit(QDate(QDate.currentDate().year(), 1, 1))
        self.f_date_from.setCalendarPopup(True)
        self.f_date_from.setDisplayFormat("dd.MM.yyyy")
        self.f_date_to = QDateEdit(QDate.currentDate())
        self.f_date_to.setCalendarPopup(True)
        self.f_date_to.setDisplayFormat("dd.MM.yyyy")
        pg_grid.addWidget(QLabel("Дата с"),  0, 0)
        pg_grid.addWidget(self.f_date_from,   0, 1)
        pg_grid.addWidget(QLabel("Дата по"), 1, 0)
        pg_grid.addWidget(self.f_date_to,     1, 1)
        pg_grid.setRowStretch(2, 1)
        self.period_stack.addWidget(page_single)

        page_batch = QWidget()
        pb_grid = QGridLayout(page_batch)
        pb_grid.setSpacing(8)
        pb_grid.setContentsMargins(0, 0, 0, 0)
        self.f_batch_from = QDateEdit(QDate(QDate.currentDate().year(),
                                            QDate.currentDate().month(), 1))
        self.f_batch_from.setCalendarPopup(True)
        self.f_batch_from.setDisplayFormat("dd.MM.yyyy")
        self.f_batch_to = QDateEdit(QDate.currentDate())
        self.f_batch_to.setCalendarPopup(True)
        self.f_batch_to.setDisplayFormat("dd.MM.yyyy")
        self.f_lookback = QSpinBox()
        self.f_lookback.setRange(1, 365)
        self.f_lookback.setValue(60)
        self.f_lookback.setSuffix(" дней")
        self.f_lookback.setStyleSheet(
            "QSpinBox { background:#1a2035; color:#e2e8f0; border:1px solid #2d3748;"
            " border-radius:6px; padding:6px 10px; }"
            "QSpinBox::up-button, QSpinBox::down-button { width:18px; }")
        pb_grid.addWidget(QLabel("Файлы с"),    0, 0)
        pb_grid.addWidget(self.f_batch_from,     0, 1)
        pb_grid.addWidget(QLabel("по"),          1, 0)
        pb_grid.addWidget(self.f_batch_to,        1, 1)
        pb_grid.addWidget(QLabel("Окно"),        2, 0)
        pb_grid.addWidget(self.f_lookback,        2, 1)
        self.period_stack.addWidget(page_batch)

        period_outer_layout.addWidget(self.period_stack)
        settings_row.addWidget(period_outer, 2)

        self.chk_single.toggled.connect(self._on_mode_toggle)
        self.chk_batch.toggled.connect(self._on_mode_toggle)

        btn_box = QGroupBox("⚡ Действия")
        btn_col = QVBoxLayout(btn_box)
        btn_col.setSpacing(10)

        self.run_btn = QPushButton("🚀  Запустить расчёт")
        self.run_btn.setObjectName("runBtn")
        self.run_btn.setMinimumHeight(48)
        self.run_btn.clicked.connect(self.start_calculation)

        self.export_btn = QPushButton("💾  Выгрузить в Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setMinimumHeight(40)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_excel)

        self.batch_run_btn = QPushButton("📦  Запустить пакет")
        self.batch_run_btn.setObjectName("runBtn")
        self.batch_run_btn.setMinimumHeight(48)
        self.batch_run_btn.setVisible(False)
        self.batch_run_btn.clicked.connect(self.start_batch)

        self.abort_btn = QPushButton("⏹  Остановить")
        self.abort_btn.setObjectName("exportBtn")
        self.abort_btn.setMinimumHeight(36)
        self.abort_btn.setVisible(False)
        self.abort_btn.clicked.connect(self.abort_batch)

        btn_col.addWidget(self.run_btn)
        btn_col.addWidget(self.export_btn)
        btn_col.addWidget(self.batch_run_btn)
        btn_col.addWidget(self.abort_btn)
        btn_col.addStretch()
        settings_row.addWidget(btn_box, 2)
        root.addLayout(settings_row)

        prog_row = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.status_lbl = QLabel("Готов к работе")
        self.status_lbl.setStyleSheet("color: #718096; font-size: 12px;")
        prog_row.addWidget(self.progress_bar, 4)
        prog_row.addWidget(self.status_lbl,    1)
        root.addLayout(prog_row)

        stats_row = QHBoxLayout()
        stats_row.setSpacing(10)
        self.stat_wagons   = self._stat_card("Вагонов",           "—")
        self.stat_turns    = self._stat_card("Оборотов",          "—")
        self.stat_avg_turn = self._stat_card("Ср. оборот, сут",   "—")
        self.stat_avg_load = self._stat_card("Ср. гружёный, сут", "—")
        for w in [self.stat_wagons, self.stat_turns, self.stat_avg_turn, self.stat_avg_load]:
            stats_row.addWidget(w)
        root.addLayout(stats_row)

        filter_box = QGroupBox("🔍 Фильтры отображения")
        filter_box.setStyleSheet(
            "QGroupBox { border:1px solid #2d3748; border-radius:6px; "
            "margin-top:8px; padding:8px; background:#161b27; "
            "font-weight:600; color:#90cdf4; font-size:12px; }")
        filter_row = QHBoxLayout(filter_box)
        filter_row.setSpacing(16)

        filter_row.addWidget(QLabel("В управлении:"))
        self.f_filter_uprav = QComboBox()
        self.f_filter_uprav.addItem("Все")
        self.f_filter_uprav.setMinimumWidth(180)
        self.f_filter_uprav.setStyleSheet(
            "QComboBox { background:#1a2035; color:#e2e8f0; border:1px solid #2d3748;"
            " border-radius:6px; padding:5px 10px; }"
            "QComboBox QAbstractItemView { background:#1a2035; color:#e2e8f0; "
            " selection-background-color:#2b4a7a; }")
        self.f_filter_uprav.currentTextChanged.connect(self._apply_filters)
        filter_row.addWidget(self.f_filter_uprav)

        refresh_btn = QPushButton("↻")
        refresh_btn.setToolTip("Обновить список компаний из БД")
        refresh_btn.setFixedSize(28, 28)
        refresh_btn.setStyleSheet(
            "QPushButton { background:#1a2035; color:#90cdf4; border:1px solid #2d3748;"
            " border-radius:6px; font-size:14px; font-weight:bold; }"
            "QPushButton:hover { background:#2b4a7a; }")
        refresh_btn.clicked.connect(self._load_companies)
        filter_row.addWidget(refresh_btn)

        filter_row.addWidget(QLabel("Статус:"))
        self.f_filter_status = QComboBox()
        self.f_filter_status.addItems(["Все", "Завершён", "Не завершён"])
        self.f_filter_status.setMinimumWidth(160)
        self.f_filter_status.setStyleSheet(
            "QComboBox { background:#1a2035; color:#e2e8f0; border:1px solid #2d3748;"
            " border-radius:6px; padding:5px 10px; }"
            "QComboBox QAbstractItemView { background:#1a2035; color:#e2e8f0;"
            " selection-background-color:#2b4a7a; }")
        self.f_filter_status.currentTextChanged.connect(self._apply_filters)
        filter_row.addWidget(self.f_filter_status)

        filter_row.addWidget(QLabel("Тип вагона:"))
        self.f_filter_type = MultiSelectButton("Все типы")
        self.f_filter_type.setMinimumWidth(160)
        self.f_filter_type.selectionChanged.connect(self._apply_filters)
        filter_row.addWidget(self.f_filter_type)

        filter_row.addStretch()

        self.filter_count_lbl = QLabel("")
        self.filter_count_lbl.setStyleSheet("color:#718096; font-size:11px;")
        filter_row.addWidget(self.filter_count_lbl)

        root.addWidget(filter_box)

        splitter = QSplitter(Qt.Vertical)
        tabs = QTabWidget()

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(self.table.styleSheet() +
            "QTableWidget { alternate-background-color: #1a2035; }")
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        tabs.addTab(self.table, "📊 Результаты")

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Лог выполнения...")
        tabs.addTab(self.log_box, "📋 Лог")

        splitter.addWidget(tabs)
        splitter.setSizes([600])
        root.addWidget(splitter, 1)

    def _load_companies(self):
        """Запускает фоновую загрузку списка управляющих компаний."""
        self.f_filter_uprav.clear()
        self.f_filter_uprav.addItem("Все")
        self.f_filter_uprav.setEnabled(False)
        self.f_filter_uprav.setPlaceholderText("Загрузка...")

        self._init_worker = InitWorker(self._conn_str())
        self._init_worker.companies_loaded.connect(self._on_companies_loaded)
        self._init_worker.types_loaded.connect(self._on_types_loaded)
        self._init_worker.error.connect(self._on_companies_error)
        self._init_worker.start()

    def _on_companies_loaded(self, companies):
        self.f_filter_uprav.blockSignals(True)
        self.f_filter_uprav.clear()
        self.f_filter_uprav.addItem("Все")
        for c in companies:
            self.f_filter_uprav.addItem(c)
        self.f_filter_uprav.blockSignals(False)
        self.f_filter_uprav.setEnabled(True)
        self._on_log(f"[{__import__('datetime').datetime.now().strftime('%H:%M:%S')}] "
                     f"Загружено компаний: {len(companies)}")

    def _on_types_loaded(self, types):
        self.f_filter_type.set_items(types)
        self._on_log(f"[{__import__('datetime').datetime.now().strftime('%H:%M:%S')}] "
                     f"Типов вагонов: {len(types)}")

    def _on_companies_error(self, msg):
        self.f_filter_uprav.setEnabled(True)
        self._on_log(f"Не удалось загрузить список компаний: {msg}")

    def _on_mode_toggle(self, checked):
        sender = self.sender()
        if not checked:
            return
        if sender is self.chk_single:
            self.chk_batch.setChecked(False)
            self.period_stack.setCurrentIndex(0)
            self.run_btn.setVisible(True)
            self.export_btn.setVisible(True)
            self.batch_run_btn.setVisible(False)
            self.abort_btn.setVisible(False)
        else:
            self.chk_single.setChecked(False)
            self.period_stack.setCurrentIndex(1)
            self.run_btn.setVisible(False)
            self.export_btn.setVisible(False)
            self.batch_run_btn.setVisible(True)
            self.abort_btn.setVisible(True)


    def _field(self, default="", password=False):
        f = QLineEdit(default)
        if password:
            f.setEchoMode(QLineEdit.Password)
        return f

    def _stat_card(self, label, value):
        lbl = QLabel(
            f"<b style='color:#90cdf4;font-size:20px'>{value}</b>"
            f"<br><span style='color:#718096;font-size:11px'>{label}</span>")
        lbl.setObjectName("statLabel")
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setMinimumWidth(160)
        lbl.setMinimumHeight(60)
        return lbl

    def _update_stat(self, widget, label, value):
        widget.setText(
            f"<b style='color:#90cdf4;font-size:20px'>{value}</b>"
            f"<br><span style='color:#718096;font-size:11px'>{label}</span>")

    def _conn_str(self):
        return (
            f"DRIVER={{ODBC Driver 18 for SQL Server}};"
            f"SERVER={self.f_server.text()};"
            f"DATABASE={self.f_database.text()};"
            f"UID={self.f_username.text()};"
            f"PWD={self.f_password.text()};"
            f"Encrypt=yes;TrustServerCertificate=yes;Connection Timeout=60;"
        )

    def start_calculation(self):
        date_from = self.f_date_from.date().toString("yyyyMMdd")
        date_to   = self.f_date_to.date().toString("yyyyMMdd")
        self.run_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_box.clear()
        self.table.setRowCount(0)
        self.worker = WorkerThread(self._conn_str(), date_from, date_to)
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_finished)
        self.worker.error.connect(self._on_error)
        self.worker.log.connect(self._on_log)
        self.worker.start()

    def start_batch(self):
        from datetime import date as _date
        qd_from = self.f_batch_from.date()
        qd_to   = self.f_batch_to.date()
        batch_from = _date(qd_from.year(), qd_from.month(), qd_from.day())
        batch_to   = _date(qd_to.year(),   qd_to.month(),   qd_to.day())
        if batch_to < batch_from:
            QMessageBox.warning(self, "Ошибка", "Дата 'по' должна быть >= 'с'")
            return
        lookback = self.f_lookback.value()
        total_files = (batch_to - batch_from).days + 1

        base_dir = os.path.dirname(
            sys.executable if getattr(sys, 'frozen', False)
            else os.path.abspath(__file__))
        output_dir = QFileDialog.getExistingDirectory(
            self, "Выберите папку для сохранения файлов", base_dir)
        if not output_dir:
            return

        self.batch_run_btn.setEnabled(False)
        self.abort_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.log_box.clear()
        self._on_log(f"Пакет: {total_files} файлов → {output_dir}")

        filters = self._get_active_filters()
        if filters:
            self._on_log(f"Активные фильтры для пакета: {filters}")
        self.batch_worker = BatchWorkerThread(
            self._conn_str(), batch_from, batch_to, lookback, output_dir,
            filters=filters)
        self.batch_worker.progress.connect(self._on_progress)
        self.batch_worker.log.connect(self._on_log)
        self.batch_worker.file_saved.connect(
            lambda p: self._on_log(f"💾 {os.path.basename(p)}"))
        self.batch_worker.error.connect(self._on_batch_error)
        self.batch_worker.finished.connect(self._on_batch_finished)
        self.batch_worker.start()

    def abort_batch(self):
        if hasattr(self, 'batch_worker') and self.batch_worker.isRunning():
            self.batch_worker.abort()
            self._on_log("⏹ Остановка после текущего файла...")
            self.abort_btn.setEnabled(False)

    def _on_batch_finished(self):
        self.batch_run_btn.setEnabled(True)
        self.abort_btn.setEnabled(False)
        self.status_lbl.setText("Пакет завершён!")
        self._on_log("✅ Все файлы сохранены")

    def _on_batch_error(self, msg):
        self.batch_run_btn.setEnabled(True)
        self.status_lbl.setText("Ошибка пакета!")
        QMessageBox.critical(self, "Ошибка пакета", f"Произошла ошибка:\n\n{msg[:500]}")
        self._on_log(f"ОШИБКА:\n{msg}")

    def _on_progress(self, pct, msg):
        self.progress_bar.setValue(pct)
        self.status_lbl.setText(msg)

    def _on_log(self, msg):
        self.log_box.append(msg)

    def _on_finished(self, df):
        self.result_df = df
        self.run_btn.setEnabled(True)
        if df.empty:
            self.status_lbl.setText("Оборотов не найдено")
            return
        self.export_btn.setEnabled(True)

        if 'Тип вагона' in df.columns:
            existing = set(self.f_filter_type._items)
            new_types = sorted([
                str(v) for v in df['Тип вагона'].dropna().unique()
                if str(v).strip() and str(v) != '-' and str(v) not in existing
            ])
            if new_types:
                all_types = sorted(existing | set(new_types))
                self.f_filter_type.set_items(all_types)

        self._apply_filters()

    def _get_active_filters(self):
        """Возвращает dict активных фильтров.
        Тип вагона хранится отдельно как set (мультивыбор)."""
        filters = {}
        uprav = self.f_filter_uprav.currentText()
        if uprav and uprav != "Все":
            filters['В управлении'] = uprav
        status = self.f_filter_status.currentText()
        if status and status != "Все":
            filters['Статус'] = status
        selected_types = self.f_filter_type.selected()
        if selected_types:
            filters['__types__'] = selected_types   # set — особый ключ для мультивыбора
        return filters

    def _filtered_df(self, df=None):
        """Применяет активные фильтры к df (или self.result_df).
        Возвращает отфильтрованный DataFrame — используется и для таблицы и для экспорта."""
        if df is None:
            df = self.result_df
        if df is None or df.empty:
            return df
        df = df.copy()
        for col, val in self._get_active_filters().items():
            if col == '__types__':
                if 'Тип вагона' in df.columns:
                    df = df[df['Тип вагона'].astype(str).isin(val)]
            elif col in df.columns:
                df = df[df[col].astype(str) == val]
        return df

    def _apply_filters(self):
        if self.result_df is None or self.result_df.empty:
            return
        df = self._filtered_df()
        total = len(self.result_df)

        active = self._get_active_filters()
        if active:
            parts = []
            for k, v in active.items():
                if k == '__types__':
                    parts.append(f"Тип: {', '.join(sorted(v))}")
                else:
                    parts.append(f"{k}: {v}")
            desc = '  |  '.join(parts)
            self.filter_count_lbl.setText(f"Фильтр: {desc}  →  {len(df):,} из {total:,}")
        else:
            self.filter_count_lbl.setText(f"Показано: {len(df):,} из {total:,}")

        self._fill_table(df)
        self._update_stats_from(df)

    def _update_stats_from(self, df):
        wagons   = df['Вагон'].nunique() if not df.empty else 0
        turns    = len(df)
        avg_turn = round(df['Оборот, сут'].mean(), 2) if not df.empty and 'Оборот, сут' in df else '-'
        avg_load = round(df['Гружёный рейс, сут'].mean(), 2) if not df.empty and 'Гружёный рейс, сут' in df else '-'
        self._update_stat(self.stat_wagons,   "Вагонов",           f"{wagons:,}")
        self._update_stat(self.stat_turns,    "Оборотов",          f"{turns:,}")
        self._update_stat(self.stat_avg_turn, "Ср. оборот, сут",   str(avg_turn))
        self._update_stat(self.stat_avg_load, "Ср. гружёный, сут", str(avg_load))

    def _on_error(self, msg):
        self.run_btn.setEnabled(True)
        self.status_lbl.setText("Ошибка!")
        QMessageBox.critical(self, "Ошибка", f"Произошла ошибка:\n\n{msg}")
        self._on_log(f"ОШИБКА:\n{msg}")

    def _fill_table(self, df):
        display_df = df.copy()
        date_cols = ['Прибытие на ст. погрузки', 'Отправление со ст. погрузки',
                     'Прибытие на ст. выгрузки', 'Отправление со ст. выгрузки',
                     'Прибытие на сл. погрузку']
        for col in date_cols:
            if col in display_df.columns:
                display_df[col] = pd.to_datetime(display_df[col], errors='coerce').dt.strftime('%d.%m.%Y %H:%M')
        limit = min(len(display_df), 5000)
        self.table.setColumnCount(len(display_df.columns))
        self.table.setHorizontalHeaderLabels(display_df.columns.tolist())
        self.table.setRowCount(limit)
        for r in range(limit):
            for c, col in enumerate(display_df.columns):
                val  = display_df.iloc[r, c]
                item = QTableWidgetItem("" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val))
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(r, c, item)
        self.table.resizeColumnsToContents()

    def export_excel(self):
        if self.result_df is None or self.result_df.empty:
            return
        try:
            base_dir = os.path.dirname(
                sys.executable if getattr(sys, 'frozen', False)
                else os.path.abspath(__file__))
            active = self._get_active_filters()
            filter_suffix = ''
            if 'В управлении' in active:
                safe = active['В управлении'][:20].replace(' ', '_').replace('"', '').replace('/', '-')
                filter_suffix = f'_{safe}'
            if 'Статус' in active and active['Статус'] == 'Завершён':
                filter_suffix += '_завершённые'
            elif 'Статус' in active and active['Статус'] == 'Не завершён':
                filter_suffix += '_незавершённые'
            fname = f"оборот_вагонов{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            path  = os.path.join(base_dir, fname)
            df_exp = self._filtered_df()
            if df_exp is None or df_exp.empty:
                QMessageBox.warning(self, "Нет данных",
                    "После применения фильтров нет данных для экспорта.")
                return
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df_exp = df_exp.copy()
                _clean_for_excel(df_exp).to_excel(writer, index=False, sheet_name='Оборот вагонов')
                ws = writer.sheets['Оборот вагонов']
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                thin   = Side(style='thin', color='2D3748')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                hdr_fill = PatternFill("solid", fgColor="1E3A5F")
                hdr_font = Font(bold=True, color="FFFFFF", size=11)
                alt_fill = PatternFill("solid", fgColor="F0F4F8")
                for ci, col in enumerate(df_exp.columns, 1):
                    cell = ws.cell(row=1, column=ci)
                    cell.fill = hdr_fill; cell.font = hdr_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                for ri in range(2, len(df_exp) + 2):
                    fill = alt_fill if ri % 2 == 0 else None
                    for ci in range(1, len(df_exp.columns) + 1):
                        cell = ws.cell(row=ri, column=ci)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        if fill: cell.fill = fill
                for ci, col in enumerate(df_exp.columns, 1):
                    ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(col)), 10) + 2, 35)
                ws.freeze_panes = 'A2'
                ws.row_dimensions[1].height = 40
            self._on_log(f"✅ Файл сохранён: {path}")
            msg = QMessageBox(self)
            msg.setWindowTitle("Готово!")
            msg.setText("Файл сохранён:")
            msg.setInformativeText(path)
            msg.setIcon(QMessageBox.Information)
            open_btn = msg.addButton("📂  Открыть папку", QMessageBox.ActionRole)
            msg.addButton("OK", QMessageBox.AcceptRole)
            msg.exec_()
            if msg.clickedButton() == open_btn:
                os.startfile(base_dir)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка сохранения", str(e))
            self._on_log(f"❌ Ошибка: {e}")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
