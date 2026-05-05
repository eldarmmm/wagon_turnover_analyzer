from datetime import datetime
import os

import pandas as pd
import pyodbc
from PyQt5.QtCore import QThread, pyqtSignal

from .turnover_logic import find_turnovers
from .utils import _clean_for_excel

"""Background workers for database loading, turnover calculation, and batch export."""

class WorkerThread(QThread):
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(pd.DataFrame)
    error    = pyqtSignal(str)
    log      = pyqtSignal(str)

    def __init__(self, conn_str, date_from, date_to):
        super().__init__()
        self.conn_str  = conn_str
        self.date_from = date_from
        self.date_to   = date_to

    def emit_log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log.emit(f"[{ts}] {msg}")

    def run(self):
        try:
            self.emit_log("Подключаемся к базе данных...")
            self.progress.emit(2, "Подключение к БД...")
            conn = pyodbc.connect(self.conn_str)
            self.emit_log("Подключено!")

            self.progress.emit(5, "Загрузка справочника станций...")
            self.emit_log("Загружаем станции...")
            st_df = pd.read_sql(
                "SELECT code, Code6, Name, id, CodeGroup FROM dbo.stations_reference WHERE code IS NOT NULL AND Name IS NOT NULL",
                conn)
            st_df['code'] = st_df['code'].astype(str).str.strip()
            st_df['Name'] = st_df['Name'].astype(str).str.strip()
            st_df['id']   = pd.to_numeric(st_df['id'], errors='coerce').astype('Int64')

            name_to_code = {}
            code_to_name = {}
            code_to_id   = {}
            for _, r in st_df.iterrows():
                c = r['code']; nm = r['Name'].upper(); sid = r['id']
                code_to_name[c] = r['Name']
                if pd.notna(sid):
                    code_to_id[c] = int(sid)
                existing = name_to_code.get(nm)
                if existing is None:
                    name_to_code[nm] = c
                else:
                    cur_5 = len(c) == 5
                    ex_5  = len(existing) == 5
                    if cur_5 and not ex_5:
                        name_to_code[nm] = c
                    elif not cur_5 and not ex_5 and len(c) < len(existing):
                        name_to_code[nm] = c
            code_to_group = {}
            st_df['CodeGroup'] = st_df['CodeGroup'].astype(str).str.strip()
            for _, r in st_df.iterrows():
                cg = r['CodeGroup']
                c  = r['code']
                group_key = cg if cg and cg not in ('', 'nan', 'None') else c
                code_to_group[c] = group_key
            self.emit_log(f"Станций: {len(name_to_code):,}")

            self.progress.emit(8, "Загрузка расстояний...")
            self.emit_log("Загружаем расстояния...")
            dist_df = pd.read_sql(
                "SELECT ID_ST_FROM, ID_ST_TO, DISTANCE FROM dbo.station_distances WHERE DISTANCE > 0 AND ID_ST_FROM IS NOT NULL AND ID_ST_TO IS NOT NULL",
                conn)
            dist_df['ID_ST_FROM'] = pd.to_numeric(dist_df['ID_ST_FROM'], errors='coerce').astype('Int64')
            dist_df['ID_ST_TO']   = pd.to_numeric(dist_df['ID_ST_TO'],   errors='coerce').astype('Int64')
            dist_df = dist_df.dropna(subset=['ID_ST_FROM', 'ID_ST_TO'])
            distance_dict = {}
            for _, r in dist_df.iterrows():
                fr, to = int(r['ID_ST_FROM']), int(r['ID_ST_TO'])
                distance_dict[(fr, to)] = r['DISTANCE']
                distance_dict[(to, fr)] = r['DISTANCE']
            self.emit_log(f"Пар расстояний: {len(distance_dict):,}")

            self.progress.emit(11, "Загрузка грузов...")
            self.emit_log("Загружаем грузы...")
            cg_df = pd.read_sql(
                "SELECT Code, Name AS Наименование FROM dbo.cargo_reference WHERE Code IS NOT NULL AND Name IS NOT NULL",
                conn)
            cg_df['Code'] = pd.to_numeric(cg_df['Code'], errors='coerce').astype('Int64')
            cargo_dict = dict(zip(cg_df['Code'], cg_df['Наименование']))
            self.emit_log(f"Грузов: {len(cargo_dict):,}")

            self.progress.emit(13, "Загрузка типов вагонов...")
            self.emit_log("Загружаем типы вагонов...")
            try:
                vt_df = pd.read_sql(
                    "SELECT DISTINCT CarNumber, ROsOwners_CarTypeName AS Тип FROM dbo.wagon_type_history WHERE CarNumber IS NOT NULL AND ROsOwners_CarTypeName IS NOT NULL",
                    conn)
                vt_df['CarNumber'] = vt_df['CarNumber'].astype(str).str.strip()
                vt_df = vt_df.drop_duplicates(subset=['CarNumber'])
                vagon_type_dict = dict(zip(vt_df['CarNumber'], vt_df['Тип']))
                self.emit_log(f"Типов вагонов: {len(vagon_type_dict):,}")
            except Exception as e:
                vagon_type_dict = {}
                self.emit_log(f"Типы вагонов недоступны: {e}")

            self.progress.emit(15, "Загрузка истории управления...")
            self.emit_log("Загружаем историю управления вагонами...")
            try:
                pp_df = pd.read_sql("""
                    SELECT
                        r.CAR_NUMBER AS carnumber,
                        Own.Name  AS Собственник,
                        Men.Name  AS В_управлении,
                        DD.DATE_INS AS Дата_начала,
                        ct.Name   AS Тип_вагона
                    FROM dbo.wagon_ownership_history r
                    LEFT JOIN dbo.documents_table  DD  ON DD.AID      = r.ID_DOCUMENT
                    LEFT JOIN dbo.counterparties_reference  Own ON r.ID_OWNER   = Own.ID
                    LEFT JOIN dbo.counterparties_reference  Men ON r.ID_MANAGER = Men.ID
                    LEFT JOIN dbo.wagon_type_reference ct  ON ct.Id        = r.CarTypeId
                    WHERE r.ID_OS IS NOT NULL
                """, conn)
                pp_df['carnumber']   = pp_df['carnumber'].astype(str).str.strip()
                pp_df['Дата_начала'] = pd.to_datetime(pp_df['Дата_начала'], errors='coerce')
                pp_df = pp_df.dropna(subset=['carnumber', 'Дата_начала'])
                pp_df = pp_df.sort_values(['carnumber', 'Дата_начала'])

                passport_dict = {}
                for _, r in pp_df.iterrows():
                    cn = r['carnumber']
                    passport_dict.setdefault(cn, []).append((
                        r['Дата_начала'],
                        str(r['Собственник'])    if pd.notna(r['Собственник'])    else '-',
                        str(r['В_управлении'])   if pd.notna(r['В_управлении'])   else '-',
                        str(r['Тип_вагона'])     if pd.notna(r['Тип_вагона'])     else '-',
                    ))
                self.emit_log(f"Записей управления: {len(pp_df):,} для {len(passport_dict):,} вагонов")
            except Exception as e:
                passport_dict = {}
                self.emit_log(f"История управления недоступна: {e}")

            self.progress.emit(17, "Загрузка групп клиентов...")
            self.emit_log("Загружаем группы клиентов...")
            cargroup_dict = {}
            try:
                cgrp_df = pd.read_sql("""
                    SELECT cig.CN, cig.DB, cig.DE, dcg.Name AS Группа
                    FROM dbo.wagon_client_groups cig
                    LEFT JOIN dbo.client_group_reference dcg ON cig.ID_Group = dcg.ID
                    WHERE cig.CN IS NOT NULL AND dcg.Name IS NOT NULL
                """, conn)
                cgrp_df['CN'] = cgrp_df['CN'].astype(str).str.strip()
                cgrp_df['CN_norm'] = cgrp_df['CN'].str.lstrip('0')
                cgrp_df['DB'] = pd.to_datetime(cgrp_df['DB'], errors='coerce')
                cgrp_df['DE'] = pd.to_datetime(cgrp_df['DE'], errors='coerce')
                cgrp_df = cgrp_df.dropna(subset=['DB'])
                try:
                    grp_id_df = pd.read_sql("SELECT ID, Name FROM dbo.client_group_reference WHERE ID IN (1103,1282)", conn)
                    id1103 = id1282 = None
                    for _, gr in grp_id_df.iterrows():
                        if int(gr['ID']) == 1103: id1103 = gr['Name']
                        if int(gr['ID']) == 1282: id1282 = gr['Name']
                except Exception:
                    id1103 = id1282 = None
                for _, cgrow in cgrp_df.iterrows():
                    key = cgrow['CN_norm']
                    grp = cgrow['Группа']
                    if id1282 and id1103 and grp == id1282:
                        grp = id1103
                    cargroup_dict.setdefault(key, []).append((cgrow['DB'], cgrow['DE'], grp))
                self.emit_log(f"Групп клиентов: {len(cargroup_dict):,}")
            except Exception as e:
                self.emit_log(f"Группы клиентов недоступны: {e}")

            self.progress.emit(20, "Загрузка операций из COH...")
            self.emit_log("Загружаем операции...")
            df = pd.read_sql(f"""
                SELECT
                    c.Carnumber         AS Номер_вагона,
                    c.DateOper          AS Дата_операции,
                    c.CodeOper          AS Код_операции,
                    c.CodeStation       AS Код_текущей_станции,
                    st_cur.Name         AS Станция_текущая,
                    st_cur.id           AS id_текущей_станции,
                    c.CodeDestStation   AS Код_станции_назначения,
                    st_dest.Name        AS Станция_назначения,
                    st_dest.id          AS id_станции_назначения,
                    c.DocNumber         AS Накладная,
                    c.CodeCargo         AS Код_груза,
                    c.Weight            AS Вес,
                    c.IsFull            AS IsFull
                FROM dbo.operations_table c WITH (NOLOCK)
                LEFT JOIN dbo.stations_reference st_cur  ON c.CodeStation     = st_cur.Code
                LEFT JOIN dbo.stations_reference st_dest ON c.CodeDestStation = st_dest.Code
                WHERE c.DateOper >= '{self.date_from}'
                  AND c.DateOper <= '{self.date_to}'
                  AND c.Carnumber IS NOT NULL
                ORDER BY c.Carnumber, c.DateOper
            """, conn)
            conn.close()
            self.emit_log(f"Загружено строк: {len(df):,}")

            self.progress.emit(42, "Подготовка данных...")

            def prep_df(d):
                d = d.copy()
                d['Дата_операции']         = pd.to_datetime(d['Дата_операции'])
                d['Код_операции']          = pd.to_numeric(d['Код_операции'],          errors='coerce').astype('Int64')
                d['Код_груза']             = pd.to_numeric(d['Код_груза'],             errors='coerce').astype('Int64')
                d['Вес']                   = pd.to_numeric(d['Вес'],                   errors='coerce').fillna(0)
                d['id_текущей_станции']    = pd.to_numeric(d['id_текущей_станции'],    errors='coerce').astype('Int64')
                d['id_станции_назначения'] = pd.to_numeric(d['id_станции_назначения'], errors='coerce').astype('Int64')
                d['IsFull']                = pd.to_numeric(d['IsFull'],                errors='coerce')
                d['Номер_вагона']          = d['Номер_вагона'].astype(str).str.strip()
                d['Код_текущей_станции']   = d['Код_текущей_станции'].astype(str).str.strip()
                d['Станция_текущая']       = d['Станция_текущая'].fillna('').astype(str).str.strip()
                d = d.dropna(subset=['Код_операции', 'Номер_вагона', 'IsFull'])
                d['IsFull']                = d['IsFull'].astype(int)
                mask_loaded_zero = (
                    (d['IsFull'] == 1) &
                    d['Накладная'].astype(str).str.match(r'^0+$')
                )
                d = d[~mask_loaded_zero]
                return d.sort_values(['Номер_вагона', 'Дата_операции'])

            df = prep_df(df)
            wagons = df['Номер_вагона'].nunique()
            self.emit_log(f"Уникальных вагонов: {wagons:,}")

            self.progress.emit(44, "Проверка вагонов на границе периода...")
            self.emit_log("Проверяем вагоны с гружёным состоянием на начало периода...")

            period_start_ts = pd.Timestamp(self.date_from)
            first_isfull = df.groupby('Номер_вагона')['IsFull'].first()
            loaded_at_start = first_isfull[first_isfull == 1].index.tolist()
            skip_first_trip = set()

            if loaded_at_start:
                self.emit_log(f"Вагонов гружёных на начало периода: {len(loaded_at_start):,}")
                cars_sql = ', '.join(f"'{c}'" for c in loaded_at_start)
                try:
                    conn2 = pyodbc.connect(self.conn_str)
                    lb_sql = (
                        "SELECT c.Carnumber AS Номер_вагона, c.DateOper AS Дата_операции,"
                        " c.CodeOper AS Код_операции, c.CodeStation AS Код_текущей_станции,"
                        " st_cur.Name AS Станция_текущая, st_cur.id AS id_текущей_станции,"
                        " c.CodeDestStation AS Код_станции_назначения,"
                        " st_dest.Name AS Станция_назначения, st_dest.id AS id_станции_назначения,"
                        " c.DocNumber AS Накладная, c.CodeCargo AS Код_груза,"
                        " c.Weight AS Вес, c.IsFull AS IsFull"
                        " FROM dbo.operations_table c WITH (NOLOCK)"
                        " LEFT JOIN dbo.stations_reference st_cur  ON c.CodeStation     = st_cur.Code"
                        " LEFT JOIN dbo.stations_reference st_dest ON c.CodeDestStation = st_dest.Code"
                        f" WHERE c.Carnumber IN ({cars_sql})"
                        f" AND c.DateOper >= DATEADD(day, -90, '{self.date_from}')"
                        f" AND c.DateOper <  '{self.date_from}'"
                        " AND c.Carnumber IS NOT NULL"
                        " ORDER BY c.Carnumber, c.DateOper"
                    )
                    lb_df = pd.read_sql(lb_sql, conn2)
                    conn2.close()
                    lb_df = prep_df(lb_df)

                    for car, grp in lb_df.groupby('Номер_вагона'):
                        car = str(car).strip()
                        grp = grp.sort_values('Дата_операции').reset_index(drop=True)
                        load_idx = None
                        for k in range(len(grp) - 1, 0, -1):
                            if grp['IsFull'].iloc[k] == 1 and grp['IsFull'].iloc[k-1] == 0:
                                load_idx = k
                                break
                        chunk = grp.iloc[load_idx:] if load_idx is not None else grp
                        dep = chunk[(chunk['Код_операции'] == 2) & (chunk['IsFull'] == 1)]
                        if dep.empty:
                            continue
                        dep_date = dep['Дата_операции'].max()
                        if dep_date < period_start_ts:
                            skip_first_trip.add(car)

                    self.emit_log(f"Рейсов вне периода (пропускаем): {len(skip_first_trip):,}")
                except Exception as e:
                    self.emit_log(f"Ошибка lookback-запроса: {e}")

            self.progress.emit(50, f"Расчёт оборотов для {wagons:,} вагонов...")
            self.emit_log("Начинаем расчёт оборотов...")

            all_results = []
            groups = list(df.groupby('Номер_вагона'))
            total  = len(groups)

            for idx, (wagon_num, group) in enumerate(groups):
                result = find_turnovers(
                    group, distance_dict, cargo_dict,
                    cargroup_dict, passport_dict,
                    name_to_code, code_to_name,
                    skip_first=(wagon_num in skip_first_trip),
                    log_fn=self.emit_log,
                    code_to_group=code_to_group)
                if not result.empty:
                    all_results.append(result)
                pct = 50 + int((idx + 1) / total * 45)
                self.progress.emit(pct, f"Обработано {idx+1:,} / {total:,} вагонов...")
                if idx % 10 == 0:
                    self.emit_log(f"Вагон {idx+1}/{total}: {wagon_num}")

            self.progress.emit(96, "Финальная сборка...")
            if all_results:
                result_df = pd.concat(all_results, ignore_index=True)
                if 'Тип вагона' in result_df.columns:
                    mask = result_df['Тип вагона'].isin(['-', '', None])
                    result_df.loc[mask, 'Тип вагона'] = (
                        result_df.loc[mask, 'Вагон'].map(vagon_type_dict).fillna('-'))
            else:
                result_df = pd.DataFrame()

            self.emit_log(f"Готово! Найдено {len(result_df):,} оборотов")
            self.progress.emit(100, "Готово!")
            self.finished.emit(result_df)

        except Exception as e:
            import traceback
            self.error.emit(traceback.format_exc())



class BatchWorkerThread(QThread):
    """Run the report for each date in a batch range and save one Excel file per day."""
    progress    = pyqtSignal(int, str)
    log         = pyqtSignal(str)
    file_saved  = pyqtSignal(str)           # путь к сохранённому файлу
    error       = pyqtSignal(str)
    finished    = pyqtSignal()

    def __init__(self, conn_str, batch_from, batch_to, lookback_days, output_dir, filters=None):
        super().__init__()
        self.conn_str      = conn_str
        self.batch_from    = batch_from     # datetime.date
        self.batch_to      = batch_to       # datetime.date
        self.lookback_days = lookback_days
        self.output_dir    = output_dir
        self.filters       = filters or {}  # {'В управлении': 'Исткомтранс', 'Статус': 'Завершён'}
        self._abort        = False

    def abort(self):
        self._abort = True

    def emit_log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log.emit(f"[{ts}] {msg}")

    def run(self):
        from datetime import timedelta
        try:
            dates = []
            d = self.batch_from
            while d <= self.batch_to:
                dates.append(d)
                d += timedelta(days=1)

            total = len(dates)
            self.emit_log(f"Пакетный режим: {total} файлов, окно {self.lookback_days} дней")

            for idx, end_date in enumerate(dates):
                if self._abort:
                    self.emit_log("Прервано пользователем")
                    break

                start_date = end_date - timedelta(days=self.lookback_days)
                date_from  = start_date.strftime('%Y%m%d')
                date_to    = end_date.strftime('%Y%m%d')
                date_str   = end_date.strftime('%d_%m_%Y')
                fname      = date_str + '.xlsx'
                fpath      = os.path.join(self.output_dir, fname)

                pct = int(idx / total * 95)
                self.progress.emit(pct, f"Файл {idx+1}/{total}: {end_date.strftime('%d.%m.%Y')}")
                self.emit_log(f"─── Файл {idx+1}/{total}: {end_date.strftime('%d.%m.%Y')} "
                              f"(период {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}) ───")

                result_df = self._run_single(date_from, date_to)

                if result_df is not None and not result_df.empty:
                    self._save_excel(result_df, fpath)
                    self.file_saved.emit(fpath)
                    self.emit_log(f"✅ Сохранён: {fname}  ({len(result_df):,} оборотов)")
                else:
                    self.emit_log(f"⚠️  Нет данных за {end_date.strftime('%d.%m.%Y')}, файл пропущен")

            self.progress.emit(100, "Пакет завершён!")
            self.finished.emit()

        except Exception as e:
            import traceback
            self.error.emit(traceback.format_exc())

    def _run_single(self, date_from, date_to):
        """Выполняет полный расчёт для одного периода, возвращает DataFrame."""
        w = WorkerThread(self.conn_str, date_from, date_to)
        w.log.connect(self.log)
        result_holder = [None]
        error_holder  = [None]

        def on_finished(df): result_holder[0] = df
        def on_error(msg):   error_holder[0]  = msg

        w.finished.connect(on_finished)
        w.error.connect(on_error)
        w.run()   # синхронный вызов

        if error_holder[0]:
            self.emit_log(f"Ошибка расчёта: {error_holder[0][:200]}")
            return None

        df = result_holder[0]
        if df is not None and not df.empty and self.filters:
            for col, val in self.filters.items():
                if col == '__types__':
                    if 'Тип вагона' in df.columns:
                        df = df[df['Тип вагона'].astype(str).isin(val)]
                elif col in df.columns:
                    df = df[df[col].astype(str) == val]
        return df

    def _save_excel(self, df, path):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        df_clean = _clean_for_excel(df)
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df_clean.to_excel(writer, index=False, sheet_name='Оборот вагонов')
            ws = writer.sheets['Оборот вагонов']
            thin     = Side(style='thin', color='2D3748')
            border   = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_fill = PatternFill("solid", fgColor="1E3A5F")
            hdr_font = Font(bold=True, color="FFFFFF", size=11)
            alt_fill = PatternFill("solid", fgColor="F0F4F8")
            for ci, col in enumerate(df_clean.columns, 1):
                cell = ws.cell(row=1, column=ci)
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            for ri in range(2, len(df_clean) + 2):
                fill = alt_fill if ri % 2 == 0 else None
                for ci in range(1, len(df_clean.columns) + 1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                    cell.border = border
                    if fill: cell.fill = fill
            for ci, col in enumerate(df_clean.columns, 1):
                ws.column_dimensions[get_column_letter(ci)].width = (
                    min(max(len(str(col)), 10) + 2, 35))
            ws.freeze_panes = 'A2'
            ws.row_dimensions[1].height = 40
